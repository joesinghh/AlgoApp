import pandas as pd
import numpy as np 
from openpyxl import load_workbook



def insert_data_main(data,cols):
    dataframe = pd.read_excel("order_data.xlsx")
    # print(data)
    df = pd.DataFrame([data],columns=cols)
    writer = pd.ExcelWriter("order_data.xlsx",engine='openpyxl')
    writer.book = load_workbook("order_data.xlsx")
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    last_row = writer.sheets['Sheet1'].max_row
    df.to_excel(writer,index=False,header=False,startrow=last_row,)
    writer.save()
    print("Done")

def insert_data(data,cols):
    t = Thread(target=insert_data_main,args=(data,cols))
    t.start()


if __name__=="__main__":
    
    dataframe = pd.read_excel("order_data.xlsx")
    last_num  = dataframe.values
    print(last_num)
    
