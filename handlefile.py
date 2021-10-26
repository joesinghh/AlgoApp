import pandas as pd
import numpy as np 
from openpyxl import load_workbook
from threading import Thread

def insert_data_main(data,cols,filename=".\\OrderData\\order_data.xlsx"):
    # dataframe = pd.read_excel(filename)

    df = pd.DataFrame([data],columns=cols)
    writer = pd.ExcelWriter(filename,engine='openpyxl')
    writer.book = load_workbook(filename)
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    last_row = writer.sheets['Sheet1'].max_row
    df.to_excel(writer,index=False,header=False,startrow=last_row,)
    writer.save()
    print("Done",filename)

def insert_data(data,cols,filename):
    t = Thread(target=insert_data_main,args=(data,cols,filename))
    t.start()

def fetch_data(filename):
    df = pd.read_excel(filename)
    return df.to_numpy().tolist()

def delete_data(row_index):
    book = load_workbook('.\\OrderData\\notcompleted.xlsx')
    sheet = book['Sheet1']
    sheet.delete_rows(row_index)
    book.save('.\\OrderData\\notcompleted.xlsx')

def change_data(data, new_sl, new_target):
    workbook = load_workbook('.\\OrderData\\notcompleted.xlsx')
    sheet = workbook['Sheet1']
    sheet.cell(data.index[0]+2,15).value = new_sl
    sheet.cell(data.index[0]+2,16).value = new_target
    workbook.save('.\\OrderData\\notcompleted.xlsx')
    

if __name__=="__main__":
    
    # dataframe = pd.read_excel(".\\OrderData\\order_data.xlsx")
    # last_num  = dataframe.values
    # print(last_num)
    delete_data(2)
